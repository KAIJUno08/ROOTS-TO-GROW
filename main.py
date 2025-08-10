from kivy.config import Config

Config.set('graphics', 'width', '400')
Config.set('graphics', 'height', '600')
Config.set('graphics', 'resizable', '0')

import sqlite3
import pandas as pd
import os
from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.spinner import Spinner
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.scrollview import ScrollView
from datetime import datetime
from fpdf import FPDF
from kivy.uix.screenmanager import Screen
from kivy.uix.image import Image
from kivy.graphics import Color, RoundedRectangle
from widgets_estilo import EtiquetaNegra, BotonBonito
from widgets_estilo import FileChooserBonito
from widgets_estilo import PantallaAzul
from widgets_estilo import PantallaBlanca2
from kivy.properties import BooleanProperty
from kivy.uix.widget import Widget
from kivy.uix.scrollview import ScrollView
from functools import partial
from widgets_estilo import EtiquetaTitulo






import openpyxl
import sqlite3

def procesar_excel_y_guardar_db(ruta_archivo, nombre_db='flores.db'):
    wb = openpyxl.load_workbook(ruta_archivo)
    hoja = wb.active

    datos = []
    for fila in hoja.iter_rows(min_row=10, values_only=True):
        if not fila[0]:
            continue
        datos.append(fila)

    conn = sqlite3.connect(nombre_db)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS siembras (
            codigo TEXT,
            nombre TEXT,
            luces INTEGER,
            semana TEXT,
            val1 REAL,
            val2 REAL,
            val3 REAL,
            val4 REAL,
            val5 REAL,
            val6 REAL,
            total REAL
        )
    ''')

    for fila in datos:
        fila = fila + (None,) * (11 - len(fila))
        cursor.execute('INSERT INTO siembras VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', fila)

    conn.commit()
    conn.close()
 








def crear_base_datos():
    conexion = sqlite3.connect('usuarios.db')
    cur = conexion.cursor()
    cur.execute('''
                   CREATE TABLE IF NOT EXISTS usuarios (
                       id INTEGER PRIMARY KEY AUTOINCREMENT,
                       nombre TEXT,
                       apellido TEXT,
                       usuario TEXT,
                       contraseña TEXT,
                       tipo_usuario TEXT
                   )
                   ''')

                        # Tabla de planificación
    cur.execute('''
                    CREATE TABLE IF NOT EXISTS planificacion (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    variedad TEXT,
                    tipo TEXT,
                    color TEXT,
                    noche_luces TEXT,
                    cubetas INTEGER
        )
    ''')

                    # Tabla de salidas
    cur.execute('''
                    CREATE TABLE IF NOT EXISTS salidas (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    variedad TEXT,
                    cubetas INTEGER,
                    garruchero TEXT,
                    fecha TEXT
        )
    ''')

                    # Tabla de remisiones
    cur.execute('''
                         CREATE TABLE IF NOT EXISTS remisiones (
                         id INTEGER PRIMARY KEY AUTOINCREMENT,
                         fecha TEXT,
                         garruchero TEXT,
                         observacion TEXT,
                         detalles TEXT
        )
    ''')

                        # Tabla de novedades
    cur.execute('''
                        CREATE TABLE IF NOT EXISTS novedades (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        fecha TEXT,
                        variedad TEXT,
                        descripcion TEXT,
                        remision_id INTEGER
        )
    ''')

    conexion.commit()
    conexion.close()
    

class PantallaBlanca(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        # Fondo blanco
        with self.canvas.before:
            Color(1, 1, 1, 1)  # Blanco RGBA
            self.rect_fondo = RoundedRectangle(size=self.size, pos=self.pos)

        self.bind(size=self.actualizar_fondo, pos=self.actualizar_fondo)

        # Logo superior izquierdo
        self.logo_izquierdo = Image(
            source='logo1.png',
            size_hint=(None, None),
            size=(120, 120),
            pos_hint={'top': 1, 'x': 0}
        )
        self.add_widget(self.logo_izquierdo)

        # Logo superior derecho
        self.logo_derecho = Image(
            source='logo2.png',
            size_hint=(None, None),
            size=(120, 120),
            pos_hint={'top': 1, 'right': 1}
        )
        self.add_widget(self.logo_derecho)

    def actualizar_fondo(self, *args):
        self.rect_fondo.size = self.size
        self.rect_fondo.pos = self.pos
        

class EtiquetaNegra(Label):
    def __init__(self, **kwargs):
        kwargs.setdefault('color', (0, 0, 0, 1))        # Texto negro
        kwargs.setdefault('font_size', 16)              # Tamaño por defecto legible
        kwargs.setdefault('halign', 'left')             # Alineación horizontal
        kwargs.setdefault('valign', 'middle')           # Alineación vertical
        super().__init__(**kwargs)

        # Para que el texto se ajuste a su contenedor
        self.bind(size=self._ajustar_texto)
        self.text_size = self.size

    def _ajustar_texto(self, *args):
        self.text_size = self.size

class BienvenidaScreen(PantallaAzul):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        layout = BoxLayout(orientation='vertical', padding=[30, 70, 30, 10], spacing=15)
        

        btn_ingresar = BotonBonito(text='Ingresar', font_size=24)
        btn_ingresar.bind(on_release=lambda x: setattr(self.manager, 'current','ingresar'))
        

        btn_registrar = BotonBonito(text='Registrar Usuario', font_size=24)
        btn_registrar.bind(on_release=lambda x: setattr(self.manager, 'current', 'registrar'))

        layout.add_widget(btn_ingresar)
        layout.add_widget(btn_registrar)
        self.add_widget(layout)


class IngresarScreen(PantallaBlanca2):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        
        with self.canvas.before:
            Color(1, 1, 1, 1) 


        layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)
        
        logo = Image(source='logo1.png', size_hint=(None, None), size=(250, 250))
        logo.pos_hint = {'center_x': 0.5}
        layout.add_widget(logo)


        layout.add_widget(Label(text="Iniciar Sesión", font_size=24, halign='center', valign='middle'))

        self.usuario_input = TextInput(hint_text="Usuario", multiline=False)
        self.contraseña_input = TextInput(hint_text="Contraseña", password=True, multiline=False)

        btn_validar = BotonBonito(text="Ingresar")
        btn_validar.bind(on_release=self.validar_credenciales)

        btn_volver = BotonBonito(text="Volver", on_release=lambda x: setattr(self.manager, 'current', 'bienvenida'))

        layout.add_widget(self.usuario_input)
        layout.add_widget(self.contraseña_input)
        layout.add_widget(btn_validar)
        layout.add_widget(btn_volver)

        self.add_widget(layout)

    def mostrar_popup(self, titulo, mensaje):
        contenido = BoxLayout(orientation='vertical', padding=10)
        contenido.add_widget(Label(text=mensaje))
        btn_cerrar = BotonBonito(text="Cerrar", size_hint=(1, 0.3))
        contenido.add_widget(btn_cerrar)
        popup = Popup(title=titulo, content=contenido, size_hint=(0.75, 0.4))
        btn_cerrar.bind(on_release=popup.dismiss)
        popup.open()

    def validar_credenciales(self, instance):
        usuario = self.usuario_input.text.strip()
        contraseña = self.contraseña_input.text.strip()

        if not usuario or not contraseña:
            self.mostrar_popup("Campos vacíos", "Por favor completa todos los campos.")
            return

        import sqlite3
        conexion = sqlite3.connect('usuarios.db')
        cursor = conexion.cursor()
        cursor.execute("SELECT tipo_usuario FROM usuarios WHERE usuario=? AND contraseña=?", (usuario, contraseña))
        resultado = cursor.fetchone()
        conexion.close()

        if resultado:
            tipo_usuario = resultado[0].strip().lower()
            self.mostrar_popup("¡Bienvenido!", f"Iniciaste sesión como: {tipo_usuario}")

            self.manager.usuario_activo = usuario
            self.manager.tipo_activo = tipo_usuario
            
            if tipo_usuario == "administrador":
                nombre_pantalla = f"admin_{usuario}"
                if not self.manager.has_screen(nombre_pantalla):
                    self.manager.add_widget(MenuAdministradorScreen(name=nombre_pantalla))
                self.manager.current = nombre_pantalla
            elif tipo_usuario == "despachador":
                nombre_pantalla = f"despacho_{usuario}"
                if not self.manager.has_screen(nombre_pantalla):
                    self.manager.add_widget(MenuDespachadorScreen(name=nombre_pantalla))
                self.manager.current = nombre_pantalla
            else:
                self.mostrar_popup("Acceso limitado", f"Tipo de usuario no reconocido: {tipo}")
        else:
            self.mostrar_popup("Error", "Usuario o contraseña incorrectos.")


class RegistrarScreen(PantallaBlanca):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)
        layout.add_widget(Label(text="Registrar Usuario", font_size=24, halign='center', valign='middle'))

        self.nombre = TextInput(hint_text="Nombre")
        self.apellido = TextInput(hint_text="Apellido")
        self.usuario = TextInput(hint_text="Usuario")
        self.contraseña = TextInput(hint_text="Contraseña", password=True)

        
        self.tipo_usuario = Spinner(
            text='Selecciona tipo de usuario',
            values=('administrador', 'despachador'),
            size_hint=(1, None),
            height=44
        )

        btn_guardar = BotonBonito(text="Registrar")
        btn_guardar.bind(on_release=self.registrar_usuario)

        btn_volver = BotonBonito(text="Volver", on_release=lambda x: setattr(self.manager, 'current', 'bienvenida'))

        layout.add_widget(self.nombre)
        layout.add_widget(self.apellido)
        layout.add_widget(self.usuario)
        layout.add_widget(self.contraseña)
        layout.add_widget(self.tipo_usuario)
        layout.add_widget(btn_guardar)
        layout.add_widget(btn_volver)

        self.add_widget(layout)

    def registrar_usuario(self, instance):
        nombre = self.nombre.text.strip()
        apellido = self.apellido.text.strip()
        usuario = self.usuario.text.strip()
        contraseña = self.contraseña.text.strip()
        tipo_usuario = self.tipo_usuario.text.strip()

        if not all([nombre, apellido, usuario, contraseña]) or tipo_usuario == 'Selecciona tipo de usuario':
            self.mostrar_popup("Campos incompletos", "Por favor completa todos los campos correctamente.")
            return

        conexion = sqlite3.connect('usuarios.db')
        cursor = conexion.cursor()
        cursor.execute('''
            INSERT INTO usuarios (nombre, apellido, usuario, contraseña, tipo_usuario)
            VALUES (?, ?, ?, ?, ?)
        ''', (nombre, apellido, usuario, contraseña, tipo_usuario))
        conexion.commit()
        conexion.close()

        self.mostrar_popup("Registro exitoso", f"Usuario {usuario} registrado como {tipo_usuario}")
        self.nombre.text = ""
        self.apellido.text = ""
        self.usuario.text = ""
        self.contraseña.text = ""
        self.tipo_usuario.text = "Selecciona tipo de usuario"

    def mostrar_popup(self, titulo, mensaje):
        contenido = BoxLayout(orientation='vertical', padding=10)
        contenido.add_widget(Label(text=mensaje))
        btn_cerrar = BotonBonito(text="Cerrar", size_hint=(1, 0.3))
        contenido.add_widget(btn_cerrar)
        popup = Popup(title=titulo, content=contenido, size_hint=(0.75, 0.4))
        btn_cerrar.bind(on_release=popup.dismiss)
        popup.open()
        
class MenuAdministradorScreen(PantallaBlanca):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)
        layout.add_widget(Label(text="Bienvenido al Menú del Administrador", font_size=24, halign='center', valign='middle'))

        layout.add_widget(Label(text="Menú del Administrador", font_size=26))

        btn_planificacion = BotonBonito(text="Planificación", font_size=20)
        btn_planificacion.bind(on_release=self.ir_a_planificacion)

        btn_resultados = BotonBonito(text="Ver resultados", font_size=20)
        btn_resultados.bind(on_release=self.ir_a_resultados)

        btn_control_usuarios = BotonBonito(text="Control de usuarios", font_size=20)
        btn_control_usuarios.bind(on_release=self.ir_a_control_usuarios)

        btn_logout = BotonBonito(text="Cerrar sesión", font_size=18,
                            on_release=lambda x: setattr(self.manager, 'current', 'bienvenida'))

        layout.add_widget(btn_planificacion)
        layout.add_widget(btn_resultados)
        layout.add_widget(btn_control_usuarios)
        layout.add_widget(btn_logout)

        self.add_widget(layout)

    def ir_a_planificacion(self, instance):
        if not self.manager.has_screen('planificacion'):
            self.manager.add_widget(PlanificacionScreen(name='planificacion'))
        self.manager.current = 'planificacion'

    def ir_a_resultados(self, instance):
            if not self.manager.has_screen('resultados'):
                self.manager.add_widget(ResultadosScreen(name='resultados'))
            self.manager.current = 'resultados'

    def ir_a_control_usuarios(self, instance):
        if not self.manager.has_screen('control_usuarios'):
            self.manager.add_widget(ControlUsuariosScreen(name='control_usuarios'))
        self.manager.current = 'control_usuarios'
        
class MenuDespachadorScreen(PantallaBlanca):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)
        layout.add_widget(Label(text="Bienvenido al Menú del Despachador", font_size=24, halign='center', valign='middle'))

        layout.add_widget(Label(text="Menú del Despachador", font_size=26))

        btn_despacho = BotonBonito(text=" Empezar despacho", font_size=20)
        btn_despacho.bind(on_release=self.ir_a_despacho)

        btn_planificacion = BotonBonito(text=" Ver planificación", font_size=20)
        btn_planificacion.bind(on_release=self.ir_a_ver_planificacion)

        btn_salida = BotonBonito(text=" Salida de la semana", font_size=20)
        btn_salida.bind(on_release=self.ir_a_salidas_semana)

        btn_remisiones = BotonBonito(text=" Remisiones", font_size=20)
        btn_remisiones.bind(on_release=self.ir_a_remisiones)

        btn_novedades = BotonBonito(text=" Novedades", font_size=20)
        btn_novedades.bind(on_release=self.ir_a_novedades)

        btn_logout = BotonBonito(text="Cerrar sesión", font_size=18,
                            on_release=lambda x: setattr(self.manager, 'current', 'bienvenida'))

        layout.add_widget(btn_despacho)
        layout.add_widget(btn_planificacion)
        layout.add_widget(btn_salida)
        layout.add_widget(btn_remisiones)
        layout.add_widget(btn_novedades)
        layout.add_widget(btn_logout)

        self.add_widget(layout)

    def ir_a_despacho(self, instance):
        if not self.manager.has_screen('despacho'):
            self.manager.add_widget(DespachoScreen(name='despacho'))
        self.manager.current = 'despacho'

    def ir_a_ver_planificacion(self, instance):
        if not self.manager.has_screen('ver_planificacion'):
            self.manager.add_widget(VerPlanificacionScreen(name='ver_planificacion'))
        else:
            self.manager.get_screen('ver_planificacion').mostrar_planificacion()
        self.manager.current = 'ver_planificacion'

    def ir_a_salidas_semana(self, instance):
        if not self.manager.has_screen('salidas_semana'):
            self.manager.add_widget(SalidasSemanaScreen(name='salidas_semana'))
        self.manager.current = 'salidas_semana'

    def ir_a_remisiones(self, instance):
        if not self.manager.has_screen('remisiones'):
            self.manager.add_widget(HistorialRemisionesScreen(name='remisiones'))
        else:
            self.manager.get_screen('remisiones').cargar_remisiones()
        self.manager.current = 'remisiones'

    def ir_a_novedades(self, instance):
        if not self.manager.has_screen('novedades'):
            self.manager.add_widget(NovedadesScreen(name='novedades'))
        else:
            self.manager.get_screen('novedades').cargar_novedades()
        self.manager.current = 'novedades'
            
class MenuUsuarioScreen(PantallaBlanca):
    def __init__(self, tipo_usuario, **kwargs):
        super().__init__(**kwargs)

        layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)
        layout.add_widget(Label(text="Menú de Usuario", font_size=26, halign='center', valign='middle'))

        layout.add_widget(Label(text=f"Bienvenido ({tipo_usuario})", font_size=24))

        if tipo_usuario.lower() == 'administrador':
            layout.add_widget(BotonBonito(text="Registrar nuevo usuario"))
            layout.add_widget(BotonBonito(text="Consultar registros"))
            layout.add_widget(BotonBonito(text="Gestionar roles"))

        elif tipo_usuario.lower() == 'despachador':
            layout.add_widget(BotonBonito(text="Crear solicitud de despacho"))
            layout.add_widget(BotonBonito(text="Consultar estado de pedidos"))

        else:
            layout.add_widget(Label(text="Tipo de usuario no reconocido."))

        layout.add_widget(BotonBonito(text="Cerrar sesión", on_release=lambda x: setattr(self.manager, 'current', 'bienvenida')))
        self.add_widget(layout)
        
class PlanificacionScreen(PantallaBlanca):
    mostrar_visor = BooleanProperty(False)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        
        self.layout = BoxLayout(orientation='vertical', padding=80, spacing=10)
        self.etiqueta_archivo = EtiquetaNegra(text="archivo actual: ninguno", font_size=18, halign='left')
        self.layout.add_widget(self.etiqueta_archivo)
        
        self.boton_actualizar = BotonBonito(text="Actualizar archivo", font_size=18)
        self.boton_actualizar.bind(on_release=self.mostrar_file_chooser)
        self.layout.add_widget(self.boton_actualizar)
        
        self.visor_archivo = FileChooserBonito(size_hint_y=0.6)
        self.visor_archivo.opacity = 0  
        self.visor_archivo.disabled = True
        self.visor_archivo.bind(selection=self.cargar_archivo)
        self.layout.add_widget(self.visor_archivo)
        
        self.layout.add_widget(Widget())
        
        self.boton_volver_menu = BotonBonito(text="Volver al menú principal", size_hint=(1, 0.15))
        self.boton_volver_menu.bind(on_release=self.volver_al_menu)
        self.layout.add_widget(self.boton_volver_menu)
        
        self.boton_cerrar = BotonBonito(text="Cerrar", size_hint_y=None, height=50,)
        self.boton_cerrar.bind(on_release=lambda x: setattr(self.manager, 'current', 'bienvenida'))
        self.layout.add_widget(self.boton_cerrar)
        
        
        self.add_widget(self.layout)
        
    def mostrar_file_chooser(self, *args):
        
        self.visor_archivo.opacity = 1
        self.visor_archivo.disabled = False
        
    def cargar_archivo(self, instancia, selection):
        if selection:
            nombre = selection[0].split('/')[-1]
            self.etiqueta_archivo.text = f"Archivo actual: {nombre}"
            
            self.visor_archivo.opacity = 0
            self.visor_archivo.disabled = True
            
            
            
    def volver_al_menu(self, instancia):
        tipo = getattr(self.manager, 'tipo_activo', None)
        usuario = getattr(self.manager, 'usuario_activo', None)

        if tipo == "administrador":
            self.manager.current = f"admin_{usuario}"
        elif tipo == "despachador":
            self.manager.current = f"despacho_{usuario}"
        else:
            self.manager.current = "bienvenida"
            
class ResultadosScreen(PantallaBlanca):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)
        self.layout.add_widget(Label(text="📊 Resultados de Planificación", font_size=24, halign='center', valign='middle'))

        self.resultados_label = Label(text="📊 Resultados de planificación", font_size=22)
        self.layout.add_widget(self.resultados_label)

        self.resultados_box = BoxLayout(orientation='vertical', spacing=5)
        self.layout.add_widget(self.resultados_box)

        btn_actualizar = BotonBonito(text="🔄 Actualizar resultados", size_hint=(1, 0.15))
        btn_actualizar.bind(on_release=self.mostrar_resultados)
        self.layout.add_widget(btn_actualizar)

        btn_volver = BotonBonito(text="⬅️ Volver al menú", size_hint=(1, 0.12))
        btn_volver.bind(on_release=self.volver_al_menu)
        self.layout.add_widget(btn_volver)

        self.add_widget(self.layout)
        self.mostrar_resultados(None)

    def mostrar_resultados(self, instance):
        self.resultados_box.clear_widgets()

        try:
            conexion = sqlite3.connect('usuarios.db')
            cursor = conexion.cursor()

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS salidas (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    variedad TEXT,
                    cubetas INTEGER
                )
            ''')

            cursor.execute('SELECT variedad, SUM(cubetas) FROM planificacion GROUP BY variedad')
            planificado = {fila[0]: fila[1] for fila in cursor.fetchall()}

            cursor.execute('SELECT variedad, SUM(cubetas) FROM salidas GROUP BY variedad')
            despachado = {fila[0]: fila[1] for fila in cursor.fetchall()}

            if not planificado:
                mensaje = "📭 No hay datos de planificación cargados.\nVe a la pantalla 'Planificación' y sube un archivo Excel."
                self.resultados_box.add_widget(Label(text=mensaje, font_size=18))
            else:
                for variedad in planificado:
                    total_plan = planificado.get(variedad, 0)
                    total_salio = despachado.get(variedad, 0)
                    faltan = total_plan - total_salio

                    novedad = "⏳ Pendiente"
                    if faltan < 0:
                        novedad = "⚠️ ¡Exceso!"
                    elif faltan == 0:
                        novedad = "✅ Completado"

                    texto = f"[b]{variedad}[/b] — Plan: {total_plan}, Salió: {total_salio}, Falta: {faltan}   {novedad}"
                    etiqueta = Label(text=texto, markup=True, font_size=16)
                    self.resultados_box.add_widget(etiqueta)

            conexion.close()

        except sqlite3.OperationalError:
            mensaje = "📭 Aún no se ha creado la tabla de planificación.\nCarga un archivo desde la pantalla 'Planificación'."
            self.resultados_box.add_widget(Label(text=mensaje, font_size=18))

    def volver_al_menu(self, instance):
        tipo = getattr(self.manager, 'tipo_activo', None)
        usuario = getattr(self.manager, 'usuario_activo', None)

        if tipo == "administrador":
            self.manager.current = f"admin_{usuario}"
        elif tipo == "despachador":
            self.manager.current = f"despacho_{usuario}"
        else:
            self.manager.current = "bienvenida"
            

class ControlUsuariosScreen(PantallaBlanca):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        
        conexion = sqlite3.connect('usuarios.db')
        cursor = conexion.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        print(cursor.fetchall())

        # Layout principal
        self.layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)

        titulo = Label(text='💜 Control de Usuarios', font_size=28, color=(0, 0, 0, 1))
        



        
        # Título principal
        self.layout.add_widget(EtiquetaTitulo(text="👥 Control de Usuarios"))
        self.layout.add_widget(titulo)

        # Campo de búsqueda (preparado para futura implementación)
        self.campo_busqueda = TextInput(hint_text="🔍 Buscar usuario...", size_hint_y=None, height=40)
        self.campo_busqueda.bind(text=self.filtrar_usuarios)
        self.layout.add_widget(self.campo_busqueda)

        # Área scrollable para mostrar la lista de usuarios
        self.lista_usuarios = BoxLayout(orientation='vertical', spacing=5, size_hint_y=None)
        self.lista_usuarios.bind(minimum_height=self.lista_usuarios.setter('height'))

        scroll = ScrollView(size_hint=(1, 1))
        scroll.add_widget(self.lista_usuarios)
        self.layout.add_widget(scroll)

        # Botón para volver al menú principal
        btn_volver = BotonBonito(text="⬅️ Volver al menú", size_hint_y=None, height=50)
        btn_volver.bind(on_release=self.volver_al_menu)
        self.layout.add_widget(Widget(size_hint_y=None, height=10))  # Espaciador visual
        self.layout.add_widget(btn_volver)

        self.add_widget(self.layout)
        self.mostrar_usuarios()

    

    def mostrar_usuarios(self):
        self.lista_usuarios.clear_widgets()  # Limpia la lista antes de recargar
        

    # Conexión a base de datos
        conexion = sqlite3.connect('usuarios.db')
        cursor = conexion.cursor()
        cursor.execute('SELECT usuario, tipo_usuario FROM usuarios')
        datos = cursor.fetchall()
        conexion.close()

        usuario_activo = getattr(self.manager, 'usuario_activo', None)

        for usuario, tipo in datos:
            fila = BoxLayout(orientation='horizontal', spacing=10, size_hint_y=None, height=40)
            

        # Etiquetas visibles con texto negro
            fila.add_widget(EtiquetaNegra(text=usuario, size_hint_x=0.4))
            fila.add_widget(EtiquetaNegra(text=tipo.capitalize(), size_hint_x=0.3))

        # Botón para cambiar a admin si es despachador
        if tipo == "despachador":
            btn_admin = BotonBonito(text="⇅ Hacer admin", size_hint_x=0.15)
            btn_admin.bind(on_release=partial(self.cambiar_a_admin, usuario))
            fila.add_widget(btn_admin)
        else:
            fila.add_widget(EtiquetaNegra(text="✔️", size_hint_x=0.15))

        # Botón para eliminar si no es el usuario activo
        if usuario != usuario_activo:
            btn_eliminar = BotonBonito(text="🗑️", size_hint_x=0.15)
            btn_eliminar.bind(on_release=partial(self.confirmar_eliminacion, usuario))
            fila.add_widget(btn_eliminar)
        else:
            fila.add_widget(EtiquetaNegra(text="🔒", size_hint_x=0.15))

        self.lista_usuarios.add_widget(fila)

    def cambiar_a_admin(self, usuario):
        conexion = sqlite3.connect('usuarios.db')
        cursor = conexion.cursor()
        cursor.execute('UPDATE usuarios SET tipo_usuario = "administrador" WHERE usuario = ?', (usuario,))
        conexion.commit()
        conexion.close()
        self.mostrar_usuarios()

    def confirmar_eliminacion(self, usuario):
    

        contenido = BoxLayout(orientation='vertical', spacing=10, padding=10)
        contenido.add_widget(Label(text=f"¿Eliminar al usuario '{usuario}'?"))

        botones = BoxLayout(size_hint_y=None, height=40, spacing=10)
        btn_si = Button(text="Sí")
        btn_si.bind(on_release=lambda x: (self.eliminar_usuario(usuario), popup.dismiss()))
        btn_no = Button(text="No")
        btn_no.bind(on_release=lambda x: popup.dismiss())
        botones.add_widget(btn_si)
        botones.add_widget(btn_no)

        contenido.add_widget(botones)

        popup = Popup(title="Confirmar eliminación", content=contenido, size_hint=(0.7, 0.3))
        popup.open()

    def eliminar_usuario(self, usuario):
        conexion = sqlite3.connect('usuarios.db')
        cursor = conexion.cursor()
        cursor.execute('DELETE FROM usuarios WHERE usuario = ?', (usuario,))
        conexion.commit()
        conexion.close()
        self.mostrar_usuarios()

    def volver_al_menu(self, instance):
        usuario = getattr(self.manager, 'usuario_activo', None)
        self.manager.current = f"admin_{usuario}"

    def filtrar_usuarios(self, instancia, texto):
        # Puedes agregar lógica de filtrado aquí si quieres implementar búsqueda
        # Por ahora simplemente recargamos la lista
        self.mostrar_usuarios()

class DespachoScreen(PantallaBlanca):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.registros_actuales = []

        self.layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)
        self.layout.add_widget(Label(text="📦 Despacho de Variedades", font_size=24, halign='center', valign='middle'))

        self.filtros = BoxLayout(orientation='horizontal', size_hint_y=None, height=40, spacing=5)
        self.input_variedad = TextInput(hint_text="Variedad", multiline=False)
        self.input_tipo = TextInput(hint_text="Tipo (Pompón o Spider)", multiline=False)
        self.input_color = TextInput(hint_text="Color", multiline=False)
        self.input_noche = TextInput(hint_text="Noche de luces", multiline=False)
        for campo in [self.input_variedad, self.input_tipo, self.input_color, self.input_noche]:
            self.filtros.add_widget(campo)

        self.layout.add_widget(Label(text="📦 Empezar despacho", font_size=22))
        self.layout.add_widget(self.filtros)

        btn_filtrar = BotonBonito(text="🔍 Filtrar variedades", size_hint_y=None, height=40)
        btn_filtrar.bind(on_release=self.filtrar_variedades)
        self.layout.add_widget(btn_filtrar)

        self.lista_resultados = BoxLayout(orientation='vertical', spacing=5)
        self.layout.add_widget(self.lista_resultados)

        self.input_garruchero = TextInput(hint_text="Nombre del garruchero", multiline=False, size_hint_y=None, height=40)
        self.layout.add_widget(self.input_garruchero)

        botones_finales = BoxLayout(size_hint_y=None, height=40, spacing=10)
        btn_volver = BotonBonito(text="⬅️ Volver al menú")
        btn_volver.bind(on_release=self.volver_al_menu)
        btn_remision = BotonBonito(text="📄 Hacer remisión")
        btn_remision.bind(on_release=self.ir_a_remision)
        botones_finales.add_widget(btn_volver)
        botones_finales.add_widget(btn_remision)
        self.layout.add_widget(botones_finales)

        self.add_widget(self.layout)

    def filtrar_variedades(self, instance):
        self.lista_resultados.clear_widgets()
        filtros = {
            'variedad': self.input_variedad.text.strip().lower(),
            'tipo': self.input_tipo.text.strip().lower(),
            'color': self.input_color.text.strip().lower(),
            'noche': self.input_noche.text.strip().lower()
        }

        try:
            con = sqlite3.connect('flores.db')
            cur = con.cursor()
            cur.execute("SELECT variedad, tipo, color, noche_luces, cubetas FROM planificacion")
            filas = cur.fetchall()
            con.close()
            
            for fila in datos:
                print(fila) 

            if not filas:
                self.lista_resultados.add_widget(Label(text="📭 No hay datos en la planificación.", font_size=16))
                return

            for v, t, c, n, cubetas in filas:
                if (filtros['variedad'] in v.lower() and
                    filtros['tipo'] in t.lower() and
                    filtros['color'] in c.lower() and
                    filtros['noche'] in n.lower()):

                    fila = BoxLayout(orientation='horizontal', spacing=5, size_hint_y=None, height=40)
                    fila.add_widget(Label(text=f"{v} ({t}) - {c} - Noche: {n} - {cubetas} cub.", size_hint_x=0.6))

                    input_cant = TextInput(hint_text="Cantidad", input_filter='int', multiline=False, size_hint_x=0.2)
                    btn_sacar = BotonBonito(text="➕ Sacar", size_hint_x=0.2)
                    btn_sacar.bind(on_release=lambda _, var=v, inp=input_cant: self.registrar_salida(var, inp))
                    fila.add_widget(input_cant)
                    fila.add_widget(btn_sacar)
                    self.lista_resultados.add_widget(fila)

        except sqlite3.OperationalError:
            self.lista_resultados.add_widget(Label(
                text="📭 Aún no hay planificación.\nCargue un archivo Excel primero.", font_size=16))

    def registrar_salida(self, variedad, input_cantidad):
        texto = input_cantidad.text.strip()
        garruchero = self.input_garruchero.text.strip()
        if not texto.isdigit() or not garruchero:
            return

        cantidad = int(texto)
        if cantidad <= 0:
            return

        con = sqlite3.connect('usuarios.db')
        cur = con.cursor()

        cur.execute("SELECT cubetas FROM planificacion WHERE variedad = ?", (variedad,))
        fila = cur.fetchone()

        if not fila:
            self.lista_resultados.add_widget(Label(
                text=f"⚠️ '{variedad}' no está en planificación.", font_size=14))
            con.close()
            return

        disponibles = fila[0]
        if cantidad > disponibles:
            self.lista_resultados.add_widget(Label(
                text=f"⚠️ Solo hay {disponibles} cubetas de '{variedad}'.", font_size=14))
            con.close()
            return

        cur.execute('''CREATE TABLE IF NOT EXISTS salidas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            variedad TEXT,
            cubetas INTEGER,
            garruchero TEXT,
            fecha TEXT)''')

        fecha = datetime.now().strftime("%Y-%m-%d %H:%M")
        hora = datetime.now().strftime("%H:%M")
        cur.execute("INSERT INTO salidas (variedad, cubetas, garruchero, fecha) VALUES (?, ?, ?, ?)",
                    (variedad, cantidad, garruchero, fecha))
        cur.execute("UPDATE planificacion SET cubetas = ? WHERE variedad = ?",
                    (disponibles - cantidad, variedad))

        con.commit()
        con.close()

        self.registros_actuales.append({
            'variedad': variedad,
            'cantidad': cantidad,
            'hora': hora
        })

        input_cantidad.text = ""
        self.lista_resultados.add_widget(Label(
            text=f"✅ Se despacharon {cantidad} de '{variedad}'", font_size=14))

    def ir_a_remision(self, instance):
        if not self.manager.has_screen('remision'):
            self.manager.add_widget(RemisionScreen(name='remision', registros=self.registros_actuales))
        else:
            rem = self.manager.get_screen('remision')
            rem.registros = self.registros_actuales
            rem.refrescar_resumen()
        self.manager.current = 'remision'

    def volver_al_menu(self, instance):
        usuario = getattr(self.manager, 'usuario_activo', None)
        self.manager.current = f"despacho_{usuario}"
        
class RemisionScreen(PantallaBlanca):
    def __init__(self, registros=None, **kwargs):
        super().__init__(**kwargs)
        self.registros = registros if registros else []

        self.layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)
        self.layout.add_widget(Label(text="📄 Remisión de Variedades", font_size=24, halign='center', valign='middle'))

        self.layout.add_widget(Label(text="📄 Confirmar remisión", font_size=22, size_hint_y=None, height=40))

        self.resumen_box = BoxLayout(orientation='vertical', spacing=5, size_hint_y=None)
        self.scroll = ScrollView(size_hint=(1, 0.4))
        self.scroll.add_widget(self.resumen_box)
        self.layout.add_widget(self.scroll)

        self.label_fecha = Label(text="", size_hint_y=None, height=30)
        self.layout.add_widget(self.label_fecha)

        self.observacion = TextInput(hint_text="Observación", multiline=True, size_hint_y=0.3)
        self.layout.add_widget(self.observacion)

        self.firma = TextInput(hint_text="Nombre del garruchero (firma)", multiline=False, size_hint_y=None, height=40)
        self.layout.add_widget(self.firma)

        botones = BoxLayout(size_hint_y=None, height=45, spacing=10)
        btn_mod = BotonBonito(text="🔁 Modificar")
        btn_mod.bind(on_release=self.volver_a_despacho)
        btn_guardar = BotonBonito(text="✅ Hacer remisión")
        btn_guardar.bind(on_release=self.guardar_remision)
        botones.add_widget(btn_mod)
        botones.add_widget(btn_guardar)

        self.layout.add_widget(botones)
        self.add_widget(self.layout)
        self.refrescar_resumen()

    def refrescar_resumen(self):
        self.resumen_box.clear_widgets()
        fecha = datetime.now().strftime("%Y-%m-%d — %H:%M")
        self.label_fecha.text = f"🗓️ Fecha: {fecha}"

        if not self.registros:
            self.resumen_box.add_widget(Label(text="No hay registros cargados.", font_size=16))
        else:
            for r in self.registros:
                texto = f"🌼 {r['variedad']} — {r['cantidad']} cubetas — ⏰ {r['hora']}"
                self.resumen_box.add_widget(Label(text=texto, font_size=15))

    def volver_a_despacho(self, instance):
        self.manager.current = 'despacho'

    def guardar_remision(self, instance):
        obs = self.ob
            
    def registrar_salida(self, variedad, input_cantidad):
        cant_texto = input_cantidad.text.strip()
        garruchero = self.input_garruchero.text.strip()
        if not cant_texto.isdigit() or not garruchero:
            return

        cantidad = int(cant_texto)
        if cantidad <= 0:
            return

        conexion = sqlite3.connect('usuarios.db')
        cursor = conexion.cursor()

        # Verificar existencia y cantidad actual
        cursor.execute("SELECT cubetas FROM planificacion WHERE variedad = ?", (variedad,))
        resultado = cursor.fetchone()

        if resultado is None:
            self.lista_resultados.add_widget(Label(
                text=f"⚠️ No se encontró planificación para '{variedad}'.", font_size=14))
            conexion.close()
            return

        cubetas_disponibles = resultado[0]
        if cantidad > cubetas_disponibles:
            self.lista_resultados.add_widget(Label(
                text=f"⚠️ No hay suficientes cubetas disponibles ({cubetas_disponibles} restantes).", font_size=14))
            conexion.close()
            return

        # Registrar salida
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS salidas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            variedad TEXT,
            cubetas INTEGER,
            garruchero TEXT,
            fecha TEXT
        )
    ''')

        fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M")
        cursor.execute("INSERT INTO salidas (variedad, cubetas, garruchero, fecha) VALUES (?, ?, ?, ?)",
                        (variedad, cantidad, garruchero, fecha_actual))

        # Actualizar planificación
        nueva_cantidad = cubetas_disponibles - cantidad
        cursor.execute("UPDATE planificacion SET cubetas = ? WHERE variedad = ?", 
                   (nueva_cantidad, variedad))

        conexion.commit()
        conexion.close()

        input_cantidad.text = ""
        self.lista_resultados.add_widget(Label(
            text=f"✅ {cantidad} cubetas despachadas de '{variedad}'", font_size=14))

    def volver_al_menu(self, instance):
        usuario = getattr(self.manager, 'usuario_activo', None)
        self.manager.current = f"despacho_{usuario}"

    def ir_a_remision(self, instance):
        # Pendiente: redireccionar a pantalla de remisión
        pass      
    
class VerPlanificacionScreen(PantallaBlanca):
    def cargar_datos(self):
        import sqlite3

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.cargar_datos()
        self.layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)
        self.layout.add_widget(Label(text="📊 Ver Planificación", font_size=24, halign='center', valign='middle'))

        self.layout.add_widget(Label(text="📋 Planificación actual", font_size=22, size_hint_y=None, height=40))

        self.scroll = ScrollView(size_hint=(1, 0.9))
        self.contenido = BoxLayout(orientation='vertical', size_hint_y=None, spacing=8, padding=10)
        self.contenido.bind(minimum_height=self.contenido.setter('height'))
        self.scroll.add_widget(self.contenido)

        self.layout.add_widget(self.scroll)

        btn_volver = BotonBonito(text="⬅️ Volver al menú", size_hint_y=None, height=45)
        btn_volver.bind(on_release=self.volver_al_menu)
        self.layout.add_widget(btn_volver)

        self.add_widget(self.layout)
        self.mostrar_planificacion()

    def mostrar_planificacion(self):
        self.contenido.clear_widgets()
        try:
            con = sqlite3.connect('flores.db')
            cur = con.cursor()
            cur.execute("SELECT variedad, tipo, color, noche_luces, cubetas FROM planificacion")
            filas = cur.fetchall()
            datos = cur.fetchall()
            print("Filas obtenidas:", filas)
            con.close()
            
            
            
            
            for fila in datos:
                print(fila)

            if not filas:
                self.contenido.add_widget(Label(text="📭 No hay planificación cargada.", font_size=16))
                return

            for v, t, c, n, cantidad in filas:
                texto = f"🌼 {v} ({t}) | 🎨 {c} | 🌙 {n} | 📦 {cantidad} cubetas"
                self.contenido.add_widget(Label(text=texto, font_size=15, size_hint_y=None, height=30))

        except sqlite3.OperationalError:
            self.contenido.add_widget(Label(
                text="⚠️ La tabla de planificación aún no ha sido creada.", font_size=16))

    def volver_al_menu(self, instance):
        usuario = getattr(self.manager, 'usuario_activo', None)
        self.manager.current = f"despacho_{usuario}"  
        
class SalidasSemanaScreen(PantallaBlanca):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)
        self.layout.add_widget(Label(text="📅 Salidas de la Semana", font_size=24, halign='center', valign='middle'))

        # Encabezado
        self.layout.add_widget(Label(text="📊 Resumen de salidas por semana", font_size=20, size_hint_y=None, height=40))

        contenido = BoxLayout(orientation='horizontal', spacing=10)

        # Columna izquierda: por día
        self.col_izquierda = BoxLayout(orientation='vertical', size_hint_x=0.55, spacing=10)
        self.scroll_izq = ScrollView(size_hint=(1, 1))
        self.box_dias = BoxLayout(orientation='vertical', size_hint_y=None, spacing=10, padding=5)
        self.box_dias.bind(minimum_height=self.box_dias.setter('height'))
        self.scroll_izq.add_widget(self.box_dias)
        self.col_izquierda.add_widget(Label(text="📆 Salidas por día", font_size=16, size_hint_y=None, height=30))
        self.col_izquierda.add_widget(self.scroll_izq)

        # Columna derecha: comparación
        self.col_derecha = BoxLayout(orientation='vertical', size_hint_x=0.45, spacing=10)
        self.scroll_der = ScrollView(size_hint=(1, 1))
        self.box_comparacion = BoxLayout(orientation='vertical', size_hint_y=None, spacing=10, padding=5)
        self.box_comparacion.bind(minimum_height=self.box_comparacion.setter('height'))
        self.scroll_der.add_widget(self.box_comparacion)
        self.col_derecha.add_widget(Label(text="📦 Comparación semanal", font_size=16, size_hint_y=None, height=30))
        self.col_derecha.add_widget(self.scroll_der)

        contenido.add_widget(self.col_izquierda)
        contenido.add_widget(self.col_derecha)
        self.layout.add_widget(contenido)

        # Botón para volver al menú
        btn_volver = BotonBonito(text="⬅️ Volver al menú", size_hint_y=None, height=45)
        btn_volver.bind(on_release=self.volver_al_menu)
        self.layout.add_widget(btn_volver)

        self.add_widget(self.layout)
        self.cargar_datos()

    def cargar_datos(self):
        import calendar
        hoy = datetime.today()
        semana_actual = hoy.isocalendar()[1]

        dias_de_la_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        salidas_por_dia = {dia: [] for dia in dias_de_la_semana}
        total_despachado = {}

        try:
            con = sqlite3.connect('usuarios.db')
            cur = con.cursor()
            cur.execute("SELECT variedad, cubetas, garruchero, fecha FROM salidas")
            filas = cur.fetchall()
            con.close()
        except:
            self.box_dias.add_widget(Label(text="⚠️ No se encontraron salidas registradas."))
            return

        for variedad, cubetas, garruchero, fecha_texto in filas:
            try:
                fecha = datetime.strptime(fecha_texto, "%Y-%m-%d %H:%M")
            except:
                continue

            if fecha.isocalendar()[1] != semana_actual:
                continue

            dia = dias_de_la_semana[fecha.weekday()]
            texto = f"🌼 {variedad} — {cubetas} cubetas ({garruchero})"
            salidas_por_dia[dia].append(texto)
            total_despachado[variedad] = total_despachado.get(variedad, 0) + cubetas

        for dia in dias_de_la_semana:
            self.box_dias.add_widget(Label(text=f"🗓 {dia}", bold=True, font_size=15))
            if salidas_por_dia[dia]:
                for item in salidas_por_dia[dia]:
                    self.box_dias.add_widget(Label(text=item, font_size=13))
            else:
                self.box_dias.add_widget(Label(text="(sin despachos)", font_size=12, italic=True))

        try:
            con = sqlite3.connect('usuarios.db')
            cur = con.cursor()
            cur.execute("SELECT variedad, cubetas FROM planificacion")
            plan = cur.fetchall()
            con.close()
        except:
            self.box_comparacion.add_widget(Label(text="⚠️ No se pudo cargar la planificación."))
            return

        for variedad, planificado in plan:
            enviado = total_despachado.get(variedad, 0)
            diferencia = enviado - planificado

            if diferencia < 0:
                color = '[color=ff3333]'  # rojo
                estado = f"Atrasado: {-diferencia}"
            elif diferencia > 0:
                color = '[color=3399ff]'  # azul
                estado = f"Adelantado: +{diferencia}"
            else:
                color = '[color=cccccc]'
                estado = "Completado justo"

            texto = f"{color}🌼 {variedad} | Planificado: {planificado} | Enviado: {enviado} | {estado}[/color]"
            self.box_comparacion.add_widget(Label(text=texto, markup=True, font_size=13))

    def volver_al_menu(self, instance):
        usuario = getattr(self.manager, 'usuario_activo', None)
        self.manager.current = f"despacho_{usuario}"

class HistorialRemisionesScreen(PantallaBlanca):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.remisiones = []
        self.filtradas = []
        self.indice = 0

        self.layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)
        self.layout.add_widget(Label(text="📁 Historial de Remisiones", font_size=24, halign='center', valign='middle'))

        self.titulo = Label(text="📁 Historial de remisiones", font_size=22, size_hint_y=None, height=40)
        self.layout.add_widget(self.titulo)

        # Campo de búsqueda
        buscador = BoxLayout(size_hint_y=None, height=45, spacing=10)
        self.input_fecha = TextInput(hint_text="Buscar por fecha (YYYY-MM-DD)", multiline=False)
        btn_buscar = BotonBonito(text="🔍 Buscar")
        btn_limpiar = BotonBonito(text="🔄 Limpiar")
        btn_buscar.bind(on_release=self.buscar_por_fecha)
        btn_limpiar.bind(on_release=self.limpiar_filtro)
        buscador.add_widget(self.input_fecha)
        buscador.add_widget(btn_buscar)
        buscador.add_widget(btn_limpiar)
        self.layout.add_widget(buscador)

        # Contenido remisión
        self.contenido = BoxLayout(orientation='vertical', spacing=8)
        self.label_fecha = Label(text="")
        self.label_garruchero = Label(text="")
        self.label_observacion = Label(text="")
        self.label_detalles = Label(text="", halign='left', valign='top')
        self.label_detalles.bind(size=self.label_detalles.setter('text_size'))

        for widget in [self.label_fecha, self.label_garruchero, self.label_observacion, self.label_detalles]:
            self.contenido.add_widget(widget)

        self.layout.add_widget(self.contenido)

        # Navegación
        botones_nav = BoxLayout(size_hint_y=None, height=45, spacing=10)
        btn_ant = BotonBonito(text="⬅️ Anterior")
        btn_sig = BotonBonito(text="Siguiente ➡️")
        btn_ant.bind(on_release=self.anterior)
        btn_sig.bind(on_release=self.siguiente)
        botones_nav.add_widget(btn_ant)
        botones_nav.add_widget(btn_sig)
        self.layout.add_widget(botones_nav)

        # Acciones
        botones_accion = BoxLayout(size_hint_y=None, height=45, spacing=10)
        btn_pdf = BotonBonito(text="📄 Guardar como PDF")
        btn_pdf.bind(on_release=self.guardar_pdf)
        btn_volver = BotonBonito(text="⬅️ Volver al menú")
        btn_volver.bind(on_release=self.volver_al_menu)
        botones_accion.add_widget(btn_pdf)
        botones_accion.add_widget(btn_volver)

        self.layout.add_widget(botones_accion)

        self.add_widget(self.layout)
        self.cargar_remisiones()

    def cargar_remisiones(self):
        try:
            con = sqlite3.connect("usuarios.db")
            cur = con.cursor()
            cur.execute("SELECT fecha, garruchero, observacion, detalles FROM remisiones ORDER BY id DESC")
            self.remisiones = cur.fetchall()
            con.close()
        except:
            self.remisiones = []

        self.filtradas = self.remisiones.copy()
        self.indice = 0
        self.mostrar_remision()

    def mostrar_remision(self):
        if not self.filtradas:
            self.label_fecha.text = "⚠️ No hay remisiones para mostrar."
            self.label_garruchero.text = ""
            self.label_observacion.text = ""
            self.label_detalles.text = ""
            return

        fecha, garruchero, obs, detalles = self.filtradas[self.indice]
        self.label_fecha.text = f"📅 Fecha: {fecha}"
        self.label_garruchero.text = f"✍️ Garruchero: {garruchero}"
        self.label_observacion.text = f"📝 Observación: {obs or '—'}"
        self.label_detalles.text = f"📦 Detalles:\n{detalles}"

    def anterior(self, instance):
        if self.indice > 0:
            self.indice -= 1
            self.mostrar_remision()

    def siguiente(self, instance):
        if self.indice < len(self.filtradas) - 1:
            self.indice += 1
            self.mostrar_remision()

    def buscar_por_fecha(self, instance):
        texto = self.input_fecha.text.strip()
        if not texto:
            return

        self.filtradas = [r for r in self.remisiones if r[0].startswith(texto)]
        self.indice = 0
        self.mostrar_remision()

    def limpiar_filtro(self, instance):
        self.input_fecha.text = ""
        self.filtradas = self.remisiones.copy()
        self.indice = 0
        self.mostrar_remision()

    def guardar_pdf(self, instance):
        if not self.filtradas:
            return

        fecha, garruchero, obs, detalles = self.filtradas[self.indice]

        if not os.path.exists("remisiones_exportadas"):
            os.makedirs("remisiones_exportadas")

        nombre_archivo = f"remision_{fecha.replace(':', '-').replace(' ', '_')}_{garruchero.replace(' ', '')}.pdf"
        ruta = os.path.join("remisiones_exportadas", nombre_archivo)

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, "REMISIÓN", ln=True, align="C")
        pdf.set_font("Arial", '', 12)
        pdf.ln(5)
        pdf.cell(0, 10, f"Fecha: {fecha}", ln=True)
        pdf.cell(0, 10, f"Garruchero: {garruchero}", ln=True)
        pdf.ln(5)
        pdf.multi_cell(0, 8, f"Observación:\n{obs or '—'}")
        pdf.ln(5)
        pdf.multi_cell(0, 8, f"Detalles:\n{detalles}")

        try:
            pdf.output(ruta)
            print(f"✅ PDF guardado en: {ruta}")
        except Exception as e:
            print(f"⚠️ Error al guardar PDF: {e}")

    def volver_al_menu(self, instance):
        usuario = getattr(self.manager, 'usuario_activo', None)
        self.manager.current = f"despacho_{usuario}"
        
class NovedadesScreen(PantallaBlanca):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.novedades = []

        self.layout = BoxLayout(orientation='vertical', padding=[10, 70, 10, 10], spacing=10)
        self.layout.add_widget(Label(text="🚨 Novedades Recientes", font_size=24, halign='center', valign='middle'))

        self.titulo = Label(text="🚨 Novedades recientes", font_size=22, size_hint_y=None, height=40)
        self.layout.add_widget(self.titulo)

        self.scroll = ScrollView()
        self.box_lista = BoxLayout(orientation='vertical', size_hint_y=None, spacing=8, padding=5)
        self.box_lista.bind(minimum_height=self.box_lista.setter('height'))
        self.scroll.add_widget(self.box_lista)

        self.layout.add_widget(self.scroll)

        # Contenedor para mostrar remisión relacionada
        self.remision_info = Label(text="", halign='left', valign='top', markup=True)
        self.remision_info.bind(size=self.remision_info.setter('text_size'))
        self.layout.add_widget(self.remision_info)

        # Botón volver
        btn_volver = BotonBonito(text="⬅️ Volver al menú", size_hint_y=None, height=45)
        btn_volver.bind(on_release=self.volver_al_menu)
        self.layout.add_widget(btn_volver)

        self.add_widget(self.layout)
        self.cargar_novedades()

    def cargar_novedades(self):
        self.box_lista.clear_widgets()
        self.novedades.clear()
        self.remision_info.text = ""

        try:
            con = sqlite3.connect("usuarios.db")
            cur = con.cursor()
            cur.execute("SELECT id, fecha, variedad, descripcion, remision_id FROM novedades ORDER BY fecha DESC")
            self.novedades = cur.fetchall()
            con.close()
        except:
            self.box_lista.add_widget(Label(text="⚠️ No se pudo cargar la tabla de novedades.", font_size=16))
            return

        if not self.novedades:
            self.box_lista.add_widget(Label(text="📭 No hay novedades registradas aún.", font_size=16))
            return

        for id_nov, fecha, variedad, desc, rem_id in self.novedades:
            texto = f"🗓 {fecha} | 🌼 {variedad} — {desc[:40]}..."
            btn = BotonBonito(text=texto, size_hint_y=None, height=45)
            btn.bind(on_release=lambda inst, r_id=rem_id: self.ver_remision(r_id))
            self.box_lista.add_widget(btn)

    def ver_remision(self, remision_id):
        try:
            con = sqlite3.connect("usuarios.db")
            cur = con.cursor()
            cur.execute("SELECT fecha, garruchero, observacion, detalles FROM remisiones WHERE id = ?", (remision_id,))
            rem = cur.fetchone()
            con.close()
        except:
            self.remision_info.text = "[color=ff0000]❌ No se pudo cargar la remisión.[/color]"
            return

        if not rem:
            self.remision_info.text = "[color=aaaaaa]Remisión no encontrada.[/color]"
            return

        fecha, garruchero, obs, detalles = rem
        self.remision_info.text = f"""
[b]🗂 Remisión relacionada:[/b]
📅 Fecha: {fecha}
✍️ Garruchero: {garruchero}
📝 Observación: {obs or '—'}
📦 Detalles:
{detalles}
"""

    def volver_al_menu(self, instance):
        usuario = getattr(self.manager, 'usuario_activo', None)
        self.manager.current = f"despacho_{usuario}"
        
class RootsToGrowApp(App):
    def build(self):
        crear_base_datos()
        
        
        sm = ScreenManager()
        sm.add_widget(BienvenidaScreen(name='bienvenida'))
        sm.add_widget(IngresarScreen(name='ingresar'))
        sm.add_widget(RegistrarScreen(name='registrar'))
        return sm
    
    

if __name__ == '__main__':
    RootsToGrowApp().run()